from flask import render_template
from flask import Flask, request, render_template
from flask_login import login_required
import plotly
import plotly.plotly as py
import plotly.graph_objs as go
import pandas
from pandas import ExcelWriter
from pandas import ExcelFile
import os
import time
import numpy
import json
from json import loads
import requests
import socket


from . import home

# append existing excel sheet with new dataframe
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

#append_df_to_excel('d:/temp/test.xlsx', df)
#append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)
#append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', index=False)
#append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', index=False, startrow=0)


@home.route('/')
def homepage():
    """
    Render the homepage template on the / route
    """
    return render_template('home/index.html', title="Welcome")


@home.route('/dashboard')
@login_required
def dashboard():
    """
    Render the dashboard template on the /dashboard route
    """
    return render_template('home/dashboard.html', title="Dashboard")


@home.route('/employer')
@login_required
def employer():
    """
    Render the dashboard template on the /dashboard route
    """
    return 'Test'

@home.route('/bdeconomicindicator')
@login_required
def bdeconomicindicator():
    readcsv = pandas.read_csv('economic indicators.csv', sep=',')
    bar1 = go.Bar(
        x = readcsv['Year'],
        y = readcsv['GDP'],
        name = 'GDP (In (Bil. US$) PPP)'
    )
    bar2 = go.Bar(
        x = readcsv['Year'],
        y = readcsv['GDP per capita'],
        name = 'GDP per capita (US$ PPP)'
    )
    bar3 = go.Bar(
        x = readcsv['Year'],
        y = readcsv['GDP growth'],
        name = 'GDP growth'
    )
    bar4 = go.Bar(
        x = readcsv['Year'],
        y = readcsv['Inflation rate'],
        name = 'Inflation rate (Percent)'
    )
    bar5 = go.Bar(
        x = readcsv['Year'],
        y = readcsv['Government debt'],
        name = 'Government debt (% of GDP)'
    )

    data = [bar1, bar2, bar3, bar4, bar5]
    graphJSON = json.dumps(data, cls=plotly.utils.PlotlyJSONEncoder)
    #legend = 'Indicator'
    #labels = ["January","February","March","April","May","June","July","August"]
    #values = [10,9,8,7,6,4,7,8]
    """
    Render the dashboard template on the /dashboard route
    """
    #return render_template('home/chart.html', values=values, labels=labels, legend=legend)
    return render_template('home/chart.html', graphJSON=graphJSON)

@home.route('/geolocation')
@login_required
def geolocation():
    """
    Render the dashboard template on the /dashboard route
    """

    return render_template('home/geolocation.html', title="Geolocation")

@home.route('/geolocation', methods=['POST'])
@login_required
def geolocation_post():
    """
    Render the dashboard template on the /dashboard route
    """
    headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/36.0.1985.125 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,application/json,image/apng,*/*;q=0.8',
        'Origin': 'https://ipgeolocation.io',
	'Referer': 'https://ipgeolocation.io',
        }
    text = request.form['website']
    url = 'https://api.ipgeolocation.io/ipgeo?include=hostname&ip=' + text
    res_text = loads(requests.get(url, headers=headers).text)
    domain = res_text['domain']
    hostname = res_text['hostname']
    ip = res_text['ip']
    country_name = res_text['country_name']
    isp = res_text['isp']
    calling_code = res_text['calling_code']
    zipcode = res_text['zipcode']
    df = pandas.DataFrame({'Domain': [domain],
                           'Hostname': [hostname],
                           'Ip': [ip],
                           'Country_Name': [country_name],
                           'Isp': [isp],
                           'Calling_Code': [calling_code],
                           'Zipcode': [zipcode]})

    append_df_to_excel('Domain-details.xlsx', df, sheet_name='Sheet1', index=False)
    #writer = ExcelWriter('Domain-details.xlsx') #header=None
    #writer = pd.ExcelWriter('c:/temp/test.xlsx', engine='openpyxl')
    #df.to_excel(writer,'Sheet1',index=False)
    #writer.save()
    exceltohtml = pandas.read_excel('Domain-details.xlsx', index=False)
    exceltohtml.columns = ['Domain','Hostname','Ip','Country_Name','Isp','Calling_Code','Zipcode']
    exceltohtml = exceltohtml[exceltohtml.Hostname.str.contains('Hostname') == False]
    return exceltohtml.to_html(classes="table table-striped table-hover table-bordered table-dark table-responsive table-sm")
    #return render_template('home/geolocation.html', title="Geolocation")


@home.route('/about')
@login_required
def about():
    """
    Render the dashboard template on the /dashboard route
    """
    return 'Test'

